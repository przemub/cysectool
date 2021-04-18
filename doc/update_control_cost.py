import json
import sys

from openpyxl import load_workbook

def main(xls_filename, model_filename):
    """
    Finds controls cost in an Excel file
    and update them in a JSON model file.
    """

    wb = load_workbook(filename=xls_filename, data_only=True)
    ws = wb.active

    with open(model_filename, "r") as file:
        model = json.load(file)

    all_found = True
    for i in range(14, 32):
        name = ws["A"][i].value
        cost = ws["H"][i].value
        ind_cost = ws["I"][i].value

        found = False
        for control in model["controls"]:
            try:
                index = model["controls"][control]["level_name"].index(name)
                found = True

                model["controls"][control]["cost"][index] = cost
                model["controls"][control]["ind_cost"][index] = ind_cost
            except ValueError:
                pass

        if not found:
            all_found = False
            print(f"{name} not found")
    
    if all_found:
        with open(model_filename, "w") as file:
            json.dump(model, file, indent=2)

if __name__ == "__main__":
    main(xls_filename=sys.argv[1], model_filename=sys.argv[2])

